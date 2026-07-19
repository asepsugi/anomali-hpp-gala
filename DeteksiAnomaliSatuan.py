# -*- coding: utf-8 -*-
"""
Engine deteksi KESALAHAN INPUT SATUAN pada data penjualan (versi tahan harga-usang).

Masalah versi lama: membandingkan harga faktur dengan harga jual master TERKINI,
padahal faktur lama bisa memakai harga lama -> false positive saat harga berubah.

Perbaikan:
- "Harga wajar untuk satuan yang diinput" diambil dari RIWAYAT PENJUALAN barang itu
  sendiri (median HARGAJUAL per barang+satuan dari JUAL.DBF), bukan dari master.
  -> tahan terhadap perubahan harga; harga lama yang konsisten dianggap wajar.
- "Satuan yang seharusnya" ditentukan dari STRUKTUR faktor isi (ISISAT2/ISISAT3 di
  master) x harga satuan-dasar lazim. Rasio antar-satuan stabil walau level harga berubah.

Anomali = harga jual jauh (> ambang) dari harga lazim satuan yang diinput, DAN cocok
dengan harga satuan lain (SALAH SATUAN) atau tidak cocok manapun (HARGA JANGGAL).
Barang jasa/non-stok dikecualikan.
"""
import os
import numpy as np
import pandas as pd
from dbfread import DBF

MIN_HIST = 2   # minimal transaksi barang+satuan agar harga lazim dianggap layak jadi patokan


def _load_dbf(path):
    df = pd.DataFrame(iter(DBF(path, ignore_missing_memofile=True, char_decode_errors='replace')))
    df.columns = [c.upper() for c in df.columns]
    return df


def _truthy(v):
    return str(v).strip().lower() in ('true', 't', '1', '1.0', 'y', 'yes')


def _build_master(stok):
    """kode -> satuan dasar; isi[(kode,satuan)] -> faktor konversi; set barang jasa/non-stok."""
    base, isi, excl = {}, {}, set()
    for _, r in stok.iterrows():
        k = r['KODEBRG']; nm = str(r.get('NAMABRG', '')).upper()
        if _truthy(r.get('NONSTOK')) or _truthy(r.get('SERVICE')) or 'ONGKOS' in nm or 'JASA' in nm:
            excl.add(k)
        b = str(r.get('SATUAN', '')).strip()
        base[k] = b; isi[(k, b)] = 1.0
        for sN, iN in (('SATUAN2', 'ISISAT2'), ('SATUAN3', 'ISISAT3')):
            su = str(r.get(sN, '')).strip()
            try:
                iv = float(r.get(iN))
            except Exception:
                iv = np.nan
            if su and su not in ('0', 'nan', 'None') and iv == iv and iv > 0:
                isi[(k, su)] = iv
    return base, isi, excl


def detect_unit_errors(db_folder, date_from, date_to, dev_threshold=0.60,
                       match_tol=0.25, progress=None):
    def say(m):
        if progress: progress(m)

    say("Membuka JUAL.DBF ...")
    jual = _load_dbf(os.path.join(db_folder, 'JUAL.DBF'))
    jual['TANGGAL'] = pd.to_datetime(jual['TANGGAL'], errors='coerce')

    say("Membaca master STOK.DBF ...")
    stok = _load_dbf(os.path.join(db_folder, 'STOK.DBF'))
    base, isi, excl = _build_master(stok)

    say("Menyaring periode ...")
    d = jual[(jual['TANGGAL'] >= date_from) & (jual['TANGGAL'] <= date_to)].copy()
    nf = d['NOFAKTUR'].astype(str).str.strip().str.upper()
    d = d[~nf.str.startswith(('R', 'T'))]
    d = d[~d['KODEBRG'].isin(excl)]
    d['SATUAN'] = d['SATUAN'].astype(str).str.strip()
    d['HJ'] = pd.to_numeric(d['HARGAJUAL'], errors='coerce')
    d = d[d['HJ'] > 0]

    say("Menghitung harga jual lazim dari riwayat penjualan ...")
    grp = d.groupby(['KODEBRG', 'SATUAN'])['HJ']
    lazim = grp.median().to_dict()
    cnt = grp.count().to_dict()

    # harga satuan-dasar lazim per barang (untuk membangun pembanding antar-satuan)
    def base_lazim(k):
        b = base.get(k); v = lazim.get((k, b))
        if v and cnt.get((k, b), 0) >= MIN_HIST:
            return v
        cands = [lazim[(kk, s)] / isi[(kk, s)] for (kk, s) in lazim
                 if kk == k and (kk, s) in isi and isi[(kk, s)] > 0 and cnt.get((kk, s), 0) >= MIN_HIST]
        return float(np.median(cands)) if cands else v
    bl = {k: base_lazim(k) for k in set(d['KODEBRG'])}

    say("Membandingkan tiap baris dengan harga lazim ...")
    out = []
    diperiksa = 0
    for x in d.itertuples(index=False):
        k = x.KODEBRG; sat = x.SATUAN; hj = x.HJ
        own = lazim.get((k, sat))
        if not own or own <= 0 or cnt.get((k, sat), 0) < MIN_HIST:
            continue
        diperiksa += 1
        dev = abs(hj - own) / own
        if dev <= dev_threshold:
            continue
        # tentukan satuan yang seharusnya dari struktur faktor isi
        b = bl.get(k); seharus = ''; hseharus = np.nan
        if b and b > 0:
            best = None
            for (kk, s), f in isi.items():
                if kk != k or s == sat:
                    continue
                exp = b * f
                if exp > 0 and abs(hj - exp) / exp <= match_tol:
                    if best is None or abs(hj - exp) < best[1]:
                        best = (s, abs(hj - exp), exp)
            if best:
                seharus, hseharus = best[0], round(best[2], 0)
        kategori = 'SALAH SATUAN' if seharus else 'HARGA JANGGAL (cek manual)'
        out.append({
            'TANGGAL': x.TANGGAL, 'NOFAKTUR': x.NOFAKTUR, 'KODEBRG': k, 'NAMABRG': x.NAMABRG,
            'SATUAN_DIINPUT': sat, 'BANYAK': getattr(x, 'ISISATUAN', np.nan),
            'HARGA_JUAL': round(hj, 0),
            'HRG_LAZIM_SATUAN_INI': round(own, 0),
            'SATUAN_SEHARUSNYA': seharus,
            'HRG_LAZIM_SEHARUSNYA': hseharus if hseharus == hseharus else '',
            'DEV_PCT': round(dev * 100, 0),
            'KATEGORI': kategori,
            'NAMAUSER': getattr(x, 'NAMAUSER', ''),
        })
    res = pd.DataFrame(out)
    if len(res):
        res = res.sort_values(['KATEGORI', 'TANGGAL'])
    ring = {
        'periode': (date_from, date_to),
        'diperiksa': int(diperiksa),
        'total': int(len(res)),
        'salah_satuan': int((res['KATEGORI'] == 'SALAH SATUAN').sum()) if len(res) else 0,
        'harga_janggal': int((res['KATEGORI'] != 'SALAH SATUAN').sum()) if len(res) else 0,
        'jasa_dikecualikan': len(excl),
    }
    return res, ring


def write_report(res, ring, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "Ringkasan"
    df, dt = ring['periode']
    lines = [
        ("DETEKSI KESALAHAN INPUT SATUAN", 12),
        (f"Periode: {df:%d/%m/%Y} s/d {dt:%d/%m/%Y}", 0),
        ("", 0),
        (f"Baris diperiksa (punya riwayat harga): {ring['diperiksa']:,}", 0),
        (f"SALAH SATUAN (harga cocok satuan lain): {ring['salah_satuan']:,}", 0),
        (f"HARGA JANGGAL (perlu cek manual): {ring['harga_janggal']:,}", 0),
        (f"Barang jasa/non-stok dikecualikan: {ring['jasa_dikecualikan']:,}", 0),
        ("", 0),
        ("Cara baca:", 11),
        ("SATUAN_DIINPUT       = satuan yang tercatat di faktur.", 0),
        ("BANYAK               = jumlah dalam satuan jual (spt di faktur).", 0),
        ("HARGA_JUAL           = harga jual di faktur.", 0),
        ("HRG_LAZIM_SATUAN_INI = harga jual LAZIM barang ini utk satuan yg diinput,", 0),
        ("                       diambil dari riwayat penjualan (bukan harga master).", 0),
        ("SATUAN_SEHARUSNYA    = satuan yang harganya cocok dengan HARGA_JUAL.", 0),
        ("DEV_PCT              = selisih HARGA_JUAL thd harga lazim satuan diinput (%).", 0),
        ("MERAH = SALAH SATUAN (yakin). ORANYE = HARGA JANGGAL (cek manual).", 0),
        ("", 0),
        ("Catatan: patokan harga diambil dari riwayat penjualan barang itu sendiri,", 0),
        ("sehingga harga lama yang konsisten TIDAK dianggap salah.", 0),
    ]
    for i, (t, sz) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=t)
        if sz: c.font = Font(bold=True, size=sz)
    ws.column_dimensions['A'].width = 72

    ws2 = wb.create_sheet("Kesalahan Satuan")
    if len(res) == 0:
        ws2.cell(row=1, column=1, value="Tidak ada kesalahan satuan pada periode ini.")
    else:
        heads = list(res.columns)
        hf = PatternFill("solid", fgColor="1F4E78"); hfont = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="D0D0D0"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
        for j, h in enumerate(heads, 1):
            c = ws2.cell(row=1, column=j, value=h); c.fill = hf; c.font = hfont
            c.alignment = Alignment(horizontal="center"); c.border = bd
        red = PatternFill("solid", fgColor="FFC7CE"); org = PatternFill("solid", fgColor="FFD9A0")
        ki = heads.index('KATEGORI')
        for i, row in enumerate(res.itertuples(index=False), start=2):
            vals = list(row); fill = red if vals[ki] == 'SALAH SATUAN' else org
            for j, v in enumerate(vals, 1):
                if hasattr(v, 'strftime'): v = v.strftime('%d/%m/%Y')
                c = ws2.cell(row=i, column=j, value=v); c.fill = fill; c.border = bd
        widths = {'TANGGAL':12,'NOFAKTUR':13,'KODEBRG':9,'NAMABRG':34,'SATUAN_DIINPUT':15,'BANYAK':9,
                  'HARGA_JUAL':12,'HRG_LAZIM_SATUAN_INI':20,'SATUAN_SEHARUSNYA':17,
                  'HRG_LAZIM_SEHARUSNYA':20,'DEV_PCT':9,'KATEGORI':24,'NAMAUSER':11}
        for j, h in enumerate(heads, 1):
            ws2.column_dimensions[get_column_letter(j)].width = widths.get(h, 12)
        ws2.freeze_panes = "A2"; ws2.auto_filter.ref = f"A1:{get_column_letter(len(heads))}{len(res)+1}"
    wb.save(out_path)
    return out_path
