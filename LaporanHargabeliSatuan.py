# -*- coding: utf-8 -*-
"""
Laporan data-quality: HARGABELI SALAH-SATUAN.

Menemukan transaksi penjualan di baris SATUAN TINGKAT (mis. BALL/LSN/RIM/DUS) yang kolom
HARGABELI-nya tampak diisi PER-SATUAN-DASAR (mis. per PCS/PAK), bukan total per baris.
Akibatnya HPP yang mestinya = harga_pokok_dasar x faktor jadi understated, dan laba di
laporan ter-overstate. Ini BUKAN anomali hitung engine, melainkan daftar pembenahan input
di aplikasi kasir.

Cara deteksi (read-only):
- Harga pokok per satuan dasar (REFBASE) dipelajari dari baris satuan DASAR barang itu
  (ISISATUAN==JUMLAH -> HARGABELI/JUMLAH), diambil MEDIAN sepanjang riwayat.
- Baris satuan TINGKAT SEJATI = ISISATUAN != JUMLAH, faktor master > 1, dan
  faktor tersirat (JUMLAH/ISISATUAN) == faktor master.
- Ditandai bila HARGABELI mendekati REFBASE (jadi ~1x harga pokok dasar), padahal
  semestinya ~ REFBASE x JUMLAH (total). -> HARGABELI diisi per-satuan-dasar.

Signature: laporan_hargabeli(db_folder, date_from=None, date_to=None, progress=None)
           -> (per_barang_df, detail_df, ring)
"""
import os
import sys
import numpy as np
import pandas as pd
from dbfread import DBF

PB_LO, PB_HI = 0.4, 1.6   # rentang HARGABELI/REFBASE agar dianggap "per-satuan-dasar"


def _load_dbf(path):
    df = pd.DataFrame(iter(DBF(path, ignore_missing_memofile=True, char_decode_errors='replace')))
    df.columns = [c.upper() for c in df.columns]
    return df


def _build_factor(stok):
    fac = {}
    for _, r in stok.iterrows():
        k = r['KODEBRG']
        fac[(k, str(r.get('SATUAN', '')).strip())] = 1.0
        for sn, iN in (('SATUAN2', 'ISISAT2'), ('SATUAN3', 'ISISAT3')):
            su = str(r.get(sn, '')).strip(); iv = r.get(iN)
            if su and su not in ('0', 'nan', 'None') and pd.notna(iv) and iv not in (0, 0.0):
                fac[(k, su)] = float(iv)
    return fac


def laporan_hargabeli(db_folder, date_from=None, date_to=None, progress=None):
    def say(m):
        if progress: progress(m)

    say("Membuka JUAL.DBF ...")
    j = _load_dbf(os.path.join(db_folder, 'JUAL.DBF'))
    j['TANGGAL'] = pd.to_datetime(j['TANGGAL'], errors='coerce')
    say("Membaca master STOK.DBF ...")
    fac = _build_factor(_load_dbf(os.path.join(db_folder, 'STOK.DBF')))

    say("Menyiapkan data ...")
    h = j[~j['NOFAKTUR'].astype(str).str.strip().str.upper().str.startswith(('R', 'T'))].copy()
    h['SATUAN'] = h['SATUAN'].astype(str).str.strip()
    isis = pd.to_numeric(h['ISISATUAN'], errors='coerce')
    juml = pd.to_numeric(h['JUMLAH'], errors='coerce')
    hb = pd.to_numeric(h['HARGABELI'], errors='coerce')
    h['ISIS'], h['JUML'], h['HB'] = isis, juml, hb
    h['FAK'] = [fac.get((k, s), np.nan) for k, s in zip(h['KODEBRG'], h['SATUAN'])]

    say("Menghitung harga pokok per satuan dasar (dari riwayat) ...")
    base = h[(isis == juml) & (juml > 0) & (hb > 0)].copy()
    base['PB'] = base['HB'] / base['JUML']
    refbase = base.groupby('KODEBRG')['PB'].median()
    h['REFBASE'] = h['KODEBRG'].map(refbase)

    say("Mendeteksi HARGABELI salah-satuan ...")
    impl = juml / isis.replace(0, np.nan)
    genuine_hi = (isis != juml) & (h['FAK'] > 1) & ((impl - h['FAK']).abs() <= 0.01) & (h['REFBASE'] > 0) & (hb > 0)
    per_base_like = (h['HB'] / h['REFBASE']).between(PB_LO, PB_HI)
    flag = genuine_hi & per_base_like

    f = h[flag].copy()
    # batasi PERIODE yang dilaporkan (patokan REFBASE tetap dari seluruh riwayat)
    if date_from is not None:
        f = f[f['TANGGAL'] >= date_from]
    if date_to is not None:
        f = f[f['TANGGAL'] <= date_to]

    f['BANYAK'] = f['ISIS']
    f['HRG_POKOK_DASAR'] = f['REFBASE'].round(0)
    f['HARGABELI_SEHARUSNYA'] = (f['REFBASE'] * f['JUML']).round(0)   # total benar utk baris
    f['HPP_BENAR_PER_SAT'] = (f['REFBASE'] * f['FAK']).round(0)       # per satuan jual
    f['HPP_LAYAR'] = (f['HB'] / f['ISIS']).round(0)
    f['LABA_LEBIH_CATAT'] = (f['REFBASE'] * f['JUML'] - f['HB']).round(0)

    detail = f[['TANGGAL', 'NOFAKTUR', 'KODEBRG', 'NAMABRG', 'SATUAN', 'BANYAK', 'JUML',
                'HB', 'HRG_POKOK_DASAR', 'HARGABELI_SEHARUSNYA', 'HPP_LAYAR',
                'HPP_BENAR_PER_SAT', 'LABA_LEBIH_CATAT', 'NAMAUSER']].copy()
    detail = detail.rename(columns={'JUML': 'JUMLAH_DASAR', 'HB': 'HARGABELI_TERTULIS'})
    detail = detail.sort_values('LABA_LEBIH_CATAT', ascending=False)

    per = (f.groupby(['KODEBRG', 'NAMABRG', 'SATUAN'])
             .agg(FAKTOR=('FAK', 'median'),
                  N_TRANSAKSI=('HB', 'count'),
                  HARGABELI_TERTULIS=('HB', 'median'),
                  HRG_POKOK_DASAR=('REFBASE', 'median'),
                  HPP_BENAR_PER_SAT=('HPP_BENAR_PER_SAT', 'median'),
                  LABA_LEBIH_CATAT=('LABA_LEBIH_CATAT', 'sum'))
             .reset_index().sort_values('LABA_LEBIH_CATAT', ascending=False))
    for c in ('FAKTOR', 'HARGABELI_TERTULIS', 'HRG_POKOK_DASAR'):
        per[c] = per[c].round(0)

    ring = {
        'periode': (date_from, date_to),
        'total_transaksi': int(len(detail)),
        'total_barang': int(per['KODEBRG'].nunique()) if len(per) else 0,
        'laba_lebih_catat': float(detail['LABA_LEBIH_CATAT'].clip(lower=0).sum()) if len(detail) else 0.0,
    }
    return per, detail, ring


def write_report(per, detail, ring, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "Ringkasan"
    df, dt = ring['periode']
    prd = "SELURUH DATA" if df is None or dt is None else f"{df:%d/%m/%Y} s/d {dt:%d/%m/%Y}"
    lines = [
        ("LAPORAN HARGABELI SALAH-SATUAN (pembenahan input kasir)", 12),
        (f"Periode: {prd}", 0),
        ("", 0),
        (f"Transaksi terindikasi           : {ring['total_transaksi']:,}", 0),
        (f"Barang terdampak                : {ring['total_barang']:,}", 0),
        (f"Perkiraan laba TER-OVERSTATE    : Rp {ring['laba_lebih_catat']:,.0f}", 0),
        ("", 0),
        ("Apa ini:", 11),
        ("Baris jual satuan TINGKAT (BALL/LSN/RIM/DUS) yang HARGABELI-nya diisi PER-SATUAN-", 0),
        ("DASAR (mis. per PCS/PAK), bukan total. Akibatnya HPP tercatat kekecilan -> laba", 0),
        ("di laporan jadi kebesaran. Perlu koreksi cara input HARGABELI di aplikasi kasir.", 0),
        ("", 0),
        ("Kolom:", 11),
        ("HARGABELI_TERTULIS   = HARGABELI yang tercatat (tampak per satuan dasar).", 0),
        ("HRG_POKOK_DASAR      = harga pokok per satuan dasar (dari riwayat).", 0),
        ("HARGABELI_SEHARUSNYA = perkiraan HARGABELI benar utk baris itu (dasar x JUMLAH).", 0),
        ("HPP_BENAR_PER_SAT    = harga pokok per satuan jual (dasar x faktor).", 0),
        ("LABA_LEBIH_CATAT     = perkiraan laba yang ter-overstate (KASAR, utk prioritas).", 0),
    ]
    for i, (t, sz) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=t)
        if sz: c.font = Font(bold=True, size=sz)
    ws.column_dimensions['A'].width = 82

    hf = PatternFill("solid", fgColor="1F4E78"); hfont = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D0D0D0"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)

    def sheet(name, data, widths):
        w = wb.create_sheet(name)
        if len(data) == 0:
            w.cell(row=1, column=1, value="Tidak ada temuan pada periode ini.")
            return
        heads = list(data.columns)
        for jx, hd in enumerate(heads, 1):
            c = w.cell(row=1, column=jx, value=hd); c.fill = hf; c.font = hfont
            c.alignment = Alignment(horizontal="center"); c.border = bd
        for i, row in enumerate(data.itertuples(index=False), start=2):
            for jx, v in enumerate(row, 1):
                if hasattr(v, 'strftime'): v = v.strftime('%d/%m/%Y')
                c = w.cell(row=i, column=jx, value=v); c.border = bd
        for jx, hd in enumerate(heads, 1):
            w.column_dimensions[get_column_letter(jx)].width = widths.get(hd, 14)
        w.freeze_panes = "A2"; w.auto_filter.ref = f"A1:{get_column_letter(len(heads))}{len(data)+1}"

    sheet("Per Barang", per, {'NAMABRG': 34, 'SATUAN': 8, 'FAKTOR': 8, 'N_TRANSAKSI': 12,
          'HARGABELI_TERTULIS': 18, 'HRG_POKOK_DASAR': 16, 'HPP_BENAR_PER_SAT': 18, 'LABA_LEBIH_CATAT': 18})
    sheet("Detail", detail, {'NAMABRG': 34, 'NOFAKTUR': 13, 'SATUAN': 8, 'BANYAK': 8, 'JUMLAH_DASAR': 12,
          'HARGABELI_TERTULIS': 18, 'HRG_POKOK_DASAR': 16, 'HARGABELI_SEHARUSNYA': 20, 'HPP_LAYAR': 11,
          'HPP_BENAR_PER_SAT': 18, 'LABA_LEBIH_CATAT': 18, 'NAMAUSER': 11})
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    import datetime as dt
    base = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.path.dirname(os.path.abspath(__file__))
    folder = sys.argv[1] if len(sys.argv) > 1 else os.path.join(base, 'data')
    d1 = pd.Timestamp(sys.argv[2]) if len(sys.argv) > 2 else None
    d2 = pd.Timestamp(sys.argv[3]) if len(sys.argv) > 3 else None
    per, detail, ring = laporan_hargabeli(folder, d1, d2, progress=lambda m: print(m))
    out_dir = os.path.join(base, 'output'); os.makedirs(out_dir, exist_ok=True)
    stamp = ("_" + f"{d1:%Y%m%d}-{d2:%Y%m%d}") if (d1 is not None and d2 is not None) else "_ALL"
    out = os.path.join(out_dir, f"Laporan_HargabeliSatuan{stamp}.xlsx")
    write_report(per, detail, ring, out)
    print(f"\nTransaksi: {ring['total_transaksi']:,} | Barang: {ring['total_barang']:,} | "
          f"Laba ter-overstate: Rp {ring['laba_lebih_catat']:,.0f}")
    print(f"File: {out}")
