# -*- coding: utf-8 -*-
"""
Engine deteksi anomali HPP (versi terkoreksi).

Mekanisme (terbukti dari aplikasi + verifikasi layar):
  HPP yang DITAMPILKAN aplikasi = HARGABELI / ISISATUAN  (bila ISISATUAN<>JUMLAH)
                                = HARGABELI / JUMLAH      (bila sama)
  -> nilai ini bisa SALAH bila HARGABELI tersimpan tidak konsisten
     (kadang per-satuan, kadang total), terutama pada satuan bertingkat (PCS/LSN/DUS).

Patokan "HPP yang benar":
  harga pokok per satuan dasar (median historis per barang, tahan perubahan harga)
  x faktor konversi satuan dari master STOK (ISISAT2/ISISAT3).
  Faktor & harga master diambil dari STOK.DBF.

Anomali = |HPP_layar - HPP_benar| / HPP_benar  melebihi ambang.
"""
import os
import numpy as np
import pandas as pd
from dbfread import DBF

MIN_ROWS_MEDIAN = 3   # minimal transaksi barang agar median dipakai; kurang dari ini pakai master
WINDOW_REF = '180D'   # jendela WAKTU MUNDUR utk patokan harga pokok sadar-waktu (trailing)


def _load_dbf(path):
    df = pd.DataFrame(iter(DBF(path, ignore_missing_memofile=True, char_decode_errors='replace')))
    df.columns = [c.upper() for c in df.columns]
    return df


def _build_master(stok):
    """Kembalikan dict harga_pokok_dasar[kode] dan faktor[(kode,satuan)] dari STOK.DBF."""
    basecost, factor = {}, {}
    for _, r in stok.iterrows():
        k = r['KODEBRG']
        bc = np.nan
        for c in ('HARGABELI', 'HARGARATA2'):
            v = r.get(c)
            if pd.notna(v) and v not in (0, 0.0):
                bc = float(v); break
        basecost[k] = bc
        factor[(k, str(r.get('SATUAN', '')).strip())] = 1.0
        for s, i in (('SATUAN2', 'ISISAT2'), ('SATUAN3', 'ISISAT3')):
            su = str(r.get(s, '')).strip(); iv = r.get(i)
            if su and su not in ('0', 'nan', 'None') and pd.notna(iv) and iv not in (0, 0.0):
                factor[(k, su)] = float(iv)
    return basecost, factor


def detect_anomalies(db_folder, date_from, date_to, tolerance=0.50, progress=None):
    """tolerance = ambang deviasi (0.50 = 50%). Kembalikan (anomali_df, ringkasan_dict)."""
    def say(m):
        if progress: progress(m)

    say("Membuka JUAL.DBF ...")
    jual = _load_dbf(os.path.join(db_folder, 'JUAL.DBF'))
    jual['TANGGAL'] = pd.to_datetime(jual['TANGGAL'], errors='coerce')

    say("Membaca master STOK.DBF ...")
    stok = _load_dbf(os.path.join(db_folder, 'STOK.DBF'))
    basecost, factor = _build_master(stok)

    say("Menyiapkan seluruh riwayat penjualan ...")
    # Patokan harga dipelajari dari SELURUH riwayat; periode terpilih hanya menyaring baris
    # yang dilaporkan (spt detektor satuan) -> periode pendek tak menyembunyikan anomali.
    hist = jual[~jual['NOFAKTUR'].astype(str).str.strip().str.upper().str.startswith(('R', 'T'))].copy()
    hist = hist.sort_values('TANGGAL').reset_index(drop=True)
    hist['SATUAN'] = hist['SATUAN'].astype(str).str.strip()
    isis = pd.to_numeric(hist['ISISATUAN'], errors='coerce')
    juml = pd.to_numeric(hist['JUMLAH'], errors='coerce')
    hb = pd.to_numeric(hist['HARGABELI'], errors='coerce')
    hist['FAKTOR'] = [factor.get((k, s), np.nan) for k, s in zip(hist['KODEBRG'], hist['SATUAN'])]

    say("Menghitung HPP (seperti tampilan aplikasi) ...")
    # Formula HPP layar aplikasi = HARGABELI/ISISATUAN (bila ISISATUAN<>JUMLAH) else HARGABELI/JUMLAH.
    # CATATAN (sudah diselidiki): utk baris satuan-tingkat, HARGABELI mayoritas (74%) tersimpan
    # sebagai TOTAL -> formula ini benar. Sebagian kecil (3%, mis. LABEL KOALA) tersimpan
    # per-satuan-dasar sehingga bisa jadi false positive; ini TIDAK dikoreksi lewat formula karena
    # penyimpanan HARGABELI tak konsisten & tak bisa dibedakan andal dari data.
    hist['HPP_LAYAR'] = np.where(isis != juml,
                                 np.where(isis != 0, hb / isis, np.nan),
                                 np.where(juml != 0, hb / juml, np.nan))
    hist['BASIS'] = hist['HPP_LAYAR'] / hist['FAKTOR']

    say("Menentukan HPP yang benar (patokan sadar-waktu) ...")
    # REF_DASAR = harga pokok per satuan dasar yang BERLAKU saat tanggal faktur = median BASIS
    # barang tsb dlm jendela WAKTU MUNDUR (trailing) sblm/saat tanggal itu. Tahan naik/turun
    # harga antar batch (mis. LOOSE LEAF A5: Jan 2024 batch 6.250, bukan 3.250 rata2 sepanjang masa).
    def _trail(g):
        r = g.rolling(WINDOW_REF, on='TANGGAL', min_periods=MIN_ROWS_MEDIAN)['BASIS'].median()
        return pd.Series(r.values, index=g.index)
    hist['REF_TL'] = hist.groupby('KODEBRG', group_keys=False)[['TANGGAL', 'BASIS']].apply(_trail)
    # fallback bila jendela mundur tak cukup sampel: median seluruh riwayat, lalu master
    med_all = hist.groupby('KODEBRG')['BASIS'].transform('median')
    cnt_all = hist.groupby('KODEBRG')['BASIS'].transform('count')
    master_arr = pd.Series([basecost.get(k, np.nan) for k in hist['KODEBRG']], index=hist.index)
    fb = pd.Series(np.where(cnt_all >= MIN_ROWS_MEDIAN, med_all, master_arr), index=hist.index)
    hist['REF_DASAR'] = hist['REF_TL'].where(hist['REF_TL'].notna(), fb)
    hist['HPP_BENAR'] = (hist['REF_DASAR'] * hist['FAKTOR']).round(0)
    hist['HPP_LAYAR'] = hist['HPP_LAYAR'].round(0)
    hist['DEV_PCT'] = np.where(hist['HPP_BENAR'] > 0,
                               ((hist['HPP_LAYAR'] - hist['HPP_BENAR']).abs() / hist['HPP_BENAR'] * 100).round(1),
                               np.nan)
    hist['RASIO'] = np.where(hist['HPP_BENAR'] > 0, (hist['HPP_LAYAR'] / hist['HPP_BENAR']).round(2), np.nan)
    qty_jual = np.where(hist['FAKTOR'] > 0, juml / hist['FAKTOR'], juml)
    hist['DAMPAK_RL'] = ((hist['HPP_BENAR'] - hist['HPP_LAYAR']) * qty_jual).round(0)

    # Klasifikasi sebab: apakah angka pokok sebenarnya masih bisa dipulihkan dari data,
    # atau memang harga pokoknya salah.
    ref = hist['REF_DASAR']
    tol_match = 0.15
    def _close(x):
        return (ref > 0) & x.notna() & ((x - ref).abs() / ref <= tol_match)
    juml_safe = juml.replace(0, np.nan)
    # Konsistensi kuantitas: satuan jual menyiratkan jumlah dasar = ISISATUAN x FAKTOR.
    # Bila JUMLAH tercatat LEBIH KECIL dari itu (mis. BALL faktor 25 tapi JUMLAH=1),
    # biaya pokok understated: HARGABELI/JUMLAH cocok ref hanya gara-gara JUMLAH-nya salah.
    # Baris seperti ini PERLU koreksi HPP, jadi jangan dicap "biaya benar".
    # (JUMLAH lebih BESAR dari tersirat = label satuan salah tapi jumlah dasar benar -> tetap
    #  boleh "TOTAL BENAR"; kalau biayanya beneran salah pun akan jatuh ke HARGA POKOK SALAH.)
    exp_juml = isis * hist['FAKTOR']
    undercount = hist['FAKTOR'].notna() & (exp_juml > 0) & (juml < exp_juml * 0.98 - 0.5)
    qty_ok = ~undercount
    c_total = qty_ok & _close(hb / juml_safe)  # HARGABELI = total utk JUMLAH -> total benar, satuan salah
    c_base = qty_ok & _close(hb)               # HARGABELI = harga pokok dasar tapi di baris satuan tinggi
    c_sell = qty_ok & _close(hb / hist['FAKTOR'])  # HARGABELI = per satuan jual
    hist['SEBAB'] = np.where(c_total, 'TOTAL BENAR (satuan salah)',
                     np.where(c_base | c_sell, 'NILAI POKOK BENAR (salah baris satuan)',
                              'HARGA POKOK SALAH'))

    say("Menyaring periode yang dilaporkan ...")
    d = hist[(hist['TANGGAL'] >= date_from) & (hist['TANGGAL'] <= date_to)]
    if len(d) == 0:
        return d, {'periode': (date_from, date_to), 'total_baris': 0, 'total_anomali': 0,
                   'hampir_pasti': 0, 'harga_pokok_salah': 0, 'satuan_tak_dikenal': 0}

    known = d['FAKTOR'].notna() & (d['HPP_BENAR'] > 0)
    anom = d[known & (d['DEV_PCT'] > tolerance * 100)].copy()
    anom['KEYAKINAN'] = np.where(anom['DEV_PCT'] > 90, 'HAMPIR PASTI',
                          np.where(anom['DEV_PCT'] > 70, 'TINGGI', 'SEDANG'))
    tak_dikenal = int((~d['FAKTOR'].notna()).sum())

    cols = ['TANGGAL', 'NOFAKTUR', 'KODEBRG', 'NAMABRG', 'SATUAN', 'JUMLAH',
            'HPP_LAYAR', 'HPP_BENAR', 'DEV_PCT', 'RASIO', 'HARGAJUAL', 'DAMPAK_RL', 'KEYAKINAN', 'SEBAB']
    if 'GANTI' in anom.columns: cols.append('GANTI')
    if 'NAMAUSER' in anom.columns: cols.append('NAMAUSER')
    cols = [c for c in cols if c in anom.columns]
    anom = anom.sort_values('TANGGAL')[cols]

    ring = {
        'periode': (date_from, date_to),
        'total_baris': int(known.sum()),
        'total_anomali': int(len(anom)),
        'hampir_pasti': int((anom['KEYAKINAN'] == 'HAMPIR PASTI').sum()) if len(anom) else 0,
        'harga_pokok_salah': int((anom['SEBAB'] == 'HARGA POKOK SALAH').sum()) if len(anom) else 0,
        'satuan_tak_dikenal': tak_dikenal,
    }
    return anom, ring


def write_report(anom, ringkasan, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "Ringkasan"
    df, dt = ringkasan['periode']
    lines = [
        ("LAPORAN DETEKSI ANOMALI HPP", 12),
        (f"Periode: {df:%Y-%m-%d} s/d {dt:%Y-%m-%d}", 0),
        ("", 0),
        (f"Baris penjualan diperiksa : {ringkasan['total_baris']:,}", 0),
        (f"Anomali ditemukan          : {ringkasan['total_anomali']:,}", 0),
        (f"  - HAMPIR PASTI (dev >90%): {ringkasan['hampir_pasti']:,}", 0),
        (f"  - HARGA POKOK SALAH (perlu koreksi angka): {ringkasan.get('harga_pokok_salah',0):,}", 0),
        (f"Baris satuan tak terdaftar di master (dicek terpisah): {ringkasan['satuan_tak_dikenal']:,}", 0),
        ("", 0),
        ("Cara baca:", 11),
        ("HPP_LAYAR = HPP yang MUNCUL di laporan aplikasi.", 0),
        ("HPP_BENAR = HPP seharusnya (harga pokok master x konversi satuan).", 0),
        ("RASIO     = HPP_LAYAR / HPP_BENAR. Dekat kelipatan bulat (mis. 12; 0,08)", 0),
        ("            menandakan salah konversi satuan.", 0),
        ("DAMPAK_RL = perkiraan salah hitung laba baris ini (Rp) - KASAR, jgn jadi nominal rugi.", 0),
        ("KEYAKINAN = HAMPIR PASTI / TINGGI / SEDANG.", 0),
        ("SEBAB     = TOTAL BENAR (satuan salah)  -> biaya benar, ISISATUAN salah.", 0),
        ("            NILAI POKOK BENAR (salah baris satuan) -> angka pokok ada, salah satuan.", 0),
        ("            HARGA POKOK SALAH -> biaya pokok beneran salah, PERLU koreksi angka.", 0),
        ("", 0),
        (">> Sheet 'Anomali' diurut: HARGA POKOK SALAH (perlu koreksi) di ATAS,", 0),
        ("   teks SEBAB-nya MERAH TEBAL. Dalam tiap grup, diurut paling menyimpang (DEV_PCT).", 0),
    ]
    for i, (t, sz) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=t)
        if sz: c.font = Font(bold=True, size=sz)
    ws.column_dimensions['A'].width = 78

    ws2 = wb.create_sheet("Anomali")
    if len(anom) == 0:
        ws2.cell(row=1, column=1, value="Tidak ada anomali pada periode & ambang ini.")
    else:
        # Urutan: HARGA POKOK SALAH (perlu koreksi) di atas, lalu paling menyimpang (DEV_PCT).
        # DEV_PCT dipilih, bukan DAMPAK_RL, karena baris JUMLAH-salah punya DAMPAK_RL kecil
        # (semu) sehingga kalau diurut dampak malah tenggelam.
        prio = (anom['SEBAB'] != 'HARGA POKOK SALAH').astype(int) if 'SEBAB' in anom.columns else 0
        dev = anom['DEV_PCT'] if 'DEV_PCT' in anom.columns else anom['DAMPAK_RL'].abs()
        a = (anom.assign(_p=prio, _dev=dev)
                 .sort_values(['_p', '_dev'], ascending=[True, False])
                 .drop(columns=['_p', '_dev']).reset_index(drop=True))
        heads = list(a.columns)
        hf = PatternFill("solid", fgColor="1F4E78"); hfont = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="D0D0D0"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
        for j, h in enumerate(heads, 1):
            c = ws2.cell(row=1, column=j, value=h); c.fill = hf; c.font = hfont
            c.alignment = Alignment(horizontal="center"); c.border = bd
        red = PatternFill("solid", fgColor="FFC7CE"); org = PatternFill("solid", fgColor="FFD9A0"); yel = PatternFill("solid", fgColor="FFF2CC")
        hps_font = Font(bold=True, color="9C0006")   # penanda HARGA POKOK SALAH
        ki = heads.index('KEYAKINAN') if 'KEYAKINAN' in heads else -1
        si = heads.index('SEBAB') if 'SEBAB' in heads else -1
        for i, row in enumerate(a.itertuples(index=False), start=2):
            vals = list(row)
            fill = yel
            if ki >= 0:
                fill = {'HAMPIR PASTI': red, 'TINGGI': org, 'SEDANG': yel}.get(vals[ki], yel)
            is_hps = (si >= 0 and vals[si] == 'HARGA POKOK SALAH')
            for j, v in enumerate(vals, 1):
                if hasattr(v, 'strftime'): v = v.strftime('%Y-%m-%d')
                elif isinstance(v, float): v = round(v, 2)
                c = ws2.cell(row=i, column=j, value=v); c.fill = fill; c.border = bd
                if is_hps and (j - 1) == si:
                    c.font = hps_font
        widths = {'TANGGAL':12,'NOFAKTUR':13,'KODEBRG':9,'NAMABRG':34,'SATUAN':7,'JUMLAH':8,
                  'HPP_LAYAR':11,'HPP_BENAR':11,'DEV_PCT':8,'RASIO':7,'HARGAJUAL':11,'DAMPAK_RL':13,'KEYAKINAN':13,'SEBAB':34,'GANTI':7,'NAMAUSER':11}
        for j, h in enumerate(heads, 1):
            ws2.column_dimensions[get_column_letter(j)].width = widths.get(h, 12)
        ws2.freeze_panes = "A2"; ws2.auto_filter.ref = f"A1:{get_column_letter(len(heads))}{len(a)+1}"
    wb.save(out_path)
    return out_path
